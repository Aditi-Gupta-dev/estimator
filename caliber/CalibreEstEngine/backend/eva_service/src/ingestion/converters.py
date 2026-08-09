"""Document -> Markdown conversion.

.docx/.pdf/.pptx go through `markitdown` (generic, handles these formats
well as markdown). .xlsx/.csv use a custom openpyxl-based converter instead,
because RAG_Architecture_Guide.docx requires worksheet/table-aware chunking
(sheet name + column headers prepended per chunk) that a generic converter
would flatten away — a plain markdown table loses the sheet boundary.
"""
import csv
from pathlib import Path

from markitdown import MarkItDown
from openpyxl import load_workbook

_markitdown = MarkItDown()

MARKITDOWN_EXTS = {".docx", ".pdf", ".pptx", ".doc", ".ppt"}
SHEET_AWARE_EXTS = {".xlsx", ".xls", ".csv"}
SUPPORTED_EXTS = MARKITDOWN_EXTS | SHEET_AWARE_EXTS


def convert_to_markdown(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in SHEET_AWARE_EXTS:
        return _sheet_aware_to_markdown(path)
    if ext in MARKITDOWN_EXTS:
        result = _markitdown.convert(str(path))
        return result.text_content
    raise ValueError(f"Unsupported file type for markdown conversion: {ext}")


def _sheet_aware_to_markdown(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _csv_to_markdown(path)
    return _xlsx_to_markdown(path)


def _csv_to_markdown(path: Path) -> str:
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return f"## Sheet: {path.stem}\n\n(empty)\n"
    return f"## Sheet: {path.stem}\n\n" + _rows_to_markdown_table(rows[0], rows[1:])


def _xlsx_to_markdown(path: Path) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(cell.strip() for cell in r)]
        if not rows:
            continue
        header, body = rows[0], rows[1:]
        sections.append(f"## Sheet: {sheet_name}\n\n" + _rows_to_markdown_table(header, body))
    wb.close()
    return "\n\n".join(sections) if sections else f"## Sheet: {path.stem}\n\n(empty)\n"


def _rows_to_markdown_table(header: list[str], body: list[list[str]]) -> str:
    header = [h or f"col{i+1}" for i, h in enumerate(header)]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        cells = list(row) + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(cells[: len(header)]) + " |")
    return "\n".join(lines) + "\n"
